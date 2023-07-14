from openpyxl import load_workbook
from datetime import datetime

from logic.const import MAIN_DIR, MONTH_LIST, DEPARTAMENT_NAMES_WITH_KEYS


class Calculation:

    def __init__(self) -> None:
        self.main_dir = MAIN_DIR
        self.year = datetime.now().year
        self.month = datetime.now().month

    def calculate(self, file_status: str, month: str) -> None:
        static_file = load_workbook(f'{MAIN_DIR}\Потребители\Реестр потребителей.xlsx', data_only=True).worksheets[0]
        svod = load_workbook(f'{MAIN_DIR}\Сводный баланс\{self.year}\УПП\Сводная ведомость {file_status} потребления.xlsx', data_only=True)
        svod_sheet  = svod.worksheets[month - 1]
        result = load_workbook('template\\rv.xlsx', data_only=True)
        result_sheet = result.worksheets[0]
 
        print("[INFO] Собираем данные")

        for row_svod in range(6, svod_sheet.max_row + 1):
            for row_static in range(9, static_file.max_row + 1):
                if svod_sheet.cell(row=row_svod, column=3).value == static_file.cell(row=row_static, column=4).value and static_file.cell(row=row_static, column=55).value == "Активен":
                    result_sheet.append(
                        [
                            1,  # A
                            static_file.cell(row=row_static, column=3).value,  # B
                            static_file.cell(row=row_static, column=4).value,  # C
                            static_file.cell(row=row_static, column=5).value,  # D
                            static_file.cell(row=row_static, column=6).value,  # E
                            str(static_file.cell(row=row_static, column=7).value),  # F
                            static_file.cell(row=row_static, column=8).value,  # G
                            static_file.cell(row=row_static, column=9).value,  # H
                            static_file.cell(row=row_static, column=12).value,  # I
                            static_file.cell(row=row_static, column=13).value,  # J
                            static_file.cell(row=row_static, column=14).value,  # K
                            static_file.cell(row=row_static, column=20).value,  # L
                            static_file.cell(row=row_static, column=21).value,  # M
                            static_file.cell(row=row_static, column=22).value,  # N
                            static_file.cell(row=row_static, column=23).value,  # O
                            static_file.cell(row=row_static, column=24).value,  # P
                            static_file.cell(row=row_static, column=27).value,  # Q
                            static_file.cell(row=row_static, column=28).value,  # R
                            str(static_file.cell(row=row_static, column=29).value),  # S
                            static_file.cell(row=row_static, column=30).value,  # T
                            svod_sheet.cell(row=row_svod, column=22).value,  # u
                            svod_sheet.cell(row=row_svod, column=22).value,  # v
                            "=V{0}-U{0}".format(row_svod),  # w
                            "=W{0}*T{0}".format(row_svod),  # x
                            svod_sheet.cell(row=row_svod, column=25).value,  # y
                            svod_sheet.cell(row=row_svod, column=26).value,  # z
                            "",  # aa
                            "",  # ab
                            "=X{0}+Y{0}+Z{0}+AA{0}+AB{0}".format(row_svod),  # ac
                            svod_sheet.cell(row=row_svod, column=30).value,  # ad
                            '',  # ae
                            "",  # af
                            "",  # ag
                            "",  # ah
                            svod_sheet.cell(row=row_svod, column=35).value,  # ai
                            svod_sheet.cell(row=row_svod, column=36).value,  # aj
                            svod_sheet.cell(row=row_svod, column=37).value,  # ak
                            svod_sheet.cell(row=row_svod, column=38).value,  # al
                            static_file.cell(row=row_static, column=54).value,  # am
                            "", # an
                            "", # ao
                            svod_sheet.cell(row=row_svod, column=40).value,  # ap
                            svod_sheet.cell(row=row_svod, column=41).value,  # aq
                        ]
                    )
                    break
            print(f"{row_svod, svod_sheet.max_row}")
        print("[INFO] Данные загружены")



        for departament in DEPARTAMENT_NAMES_WITH_KEYS.keys():
            departament_data = []
            for row in range(6, result_sheet.max_row +1):
                try: 
                    if result_sheet.cell(row=row, column=3).value[2:4] == departament:
                        departament_data.append([result_sheet.cell(row=row, column=i).value for i in range(1, result_sheet.max_column + 1)])
                except Exception:
                    pass
            workbook = load_workbook('./template/rv.xlsx', data_only=True)
            worksheet = workbook.worksheets[0]
            for departament_row in departament_data:
                worksheet.append(departament_row)
            for report_row in range(6, worksheet.max_row+1):
                worksheet["W{0}".format(report_row)].value = "=V{0}-U{0}".format(report_row)
                worksheet["X{0}".format(report_row)].value = "=W{0}*T{0}".format(report_row)
                worksheet["AC{0}".format(report_row)].value = "=X{0}+Y{0}+Z{0}+AA{0}+AB{0}".format(report_row)
                worksheet.cell(row=report_row, column=25).value = "=IF(ISBLANK($AL${0}),0,ROUND(($X${0}*$AL${0}),0))".format(report_row)
            
                
                if worksheet.cell(row=report_row, column=6).value == "None":
                    worksheet.cell(row=report_row, column=6).value = ""
                if worksheet.cell(row=report_row, column=19).value == "None":
                    worksheet.cell(row=report_row, column=19).value = ""
                    
            worksheet.cell(row=4, column=23).value = "=SUM(W6:W{0})".format(worksheet.max_row + 1)
            worksheet.cell(row=4, column=24).value = "=SUM(X6:X{0})".format(worksheet.max_row + 1)
            worksheet.cell(row=4, column=25).value = "=SUM(Y6:Y{0})".format(worksheet.max_row + 1)
            worksheet.cell(row=4, column=26).value = "=SUM(Z6:Z{0})".format(worksheet.max_row + 1)
            worksheet.cell(row=4, column=27).value = "=SUM(AA6:AA{0})".format(worksheet.max_row + 1)
            worksheet.cell(row=4, column=28).value = "=SUM(AB6:AB{0})".format(worksheet.max_row + 1)
            worksheet.cell(row=4, column=29).value = "=SUM(AC6:AC{0})".format(worksheet.max_row + 1)
            
            stuf_name = DEPARTAMENT_NAMES_WITH_KEYS[departament]
            mon = MONTH_LIST[month]

            workbook.save(f"{MAIN_DIR}\Сводный баланс\{self.year}\УПП\{MONTH_LIST[month]}\{stuf_name}\РВ {stuf_name} {file_status} потребления {mon} {self.year}.xlsx")
            print(f"[INFO] Файл " + f"РВ {stuf_name} {file_status} потребления {mon} {self.year}.xlsx сформирован")

