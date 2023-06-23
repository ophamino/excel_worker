import os
from datetime import datetime

from openpyxl import load_workbook
from logic.const import MAIN_DIR


class Report:

    def __init__(self) -> None:
        self.year = datetime.now().year

    def comparer(self, file_status: str, month: str):

        svod = load_workbook(f'{MAIN_DIR}\Сводный баланс\{self.year}\УПП\Сводный баланс {file_status} потребления.xlsx')
        svod_sheet = svod.worksheets[month]
        report = load_workbook('./template/report_gp.xlsx')
        report_sheet = report.worksheets[0]

        for row_svod in range(6, svod_sheet.max_row):
            data = [
                1,
                svod_sheet.cell(row=row_svod, column=2).value,  # B
                svod_sheet.cell(row=row_svod, column=6).value,  # C
                svod_sheet.cell(row=row_svod, column=8).value,  # D
                svod_sheet.cell(row=row_svod, column=18).value,  # E
                svod_sheet.cell(row=row_svod, column=19).value,  # F
                svod_sheet.cell(row=row_svod, column=20).value,  # G
                svod_sheet.cell(row=row_svod, column=21).value,  # H
                svod_sheet.cell(row=row_svod, column=22).value,  # I
                svod_sheet.cell(row=row_svod, column=23).value,  # J
                svod_sheet.cell(row=row_svod, column=24).value,  # K
                svod_sheet.cell(row=row_svod, column=25).value,  # L
                svod_sheet.cell(row=row_svod, column=26).value,
                svod_sheet.cell(row=row_svod, column=28).value,
                svod_sheet.cell(row=row_svod, column=27).value,
                svod_sheet.cell(row=row_svod, column=29).value,
                svod_sheet.cell(row=row_svod, column=30).value,


                
                svod_sheet.cell(row=row_svod, column=31).value,
            ]

            report_sheet.append(data)

        report.save(f"{MAIN_DIR}\Аналитика\Отчеты\Отчет ГП.xlsx")
