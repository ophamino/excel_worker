from openpyxl import load_workbook
import os
from datetime import datetime

from logic.const import MAIN_DIR, MONTH_LIST, DEPARTAMENT_NAMES_WITH_KEYS


def comparer(month: str) -> None:
        
        static_path = f'{MAIN_DIR}\Сводный баланс\{datetime.now().year}\Бику'

        if not static_path in os.listdir(static_path):
            svod = load_workbook('template/bicu.xlsx')
            svod.save(f"{static_path}\Сводная ведомость БИКУ.xlsx")
        svod = load_workbook(f"{static_path}\Сводная ведомость БИКУ.xlsx")
        svod_sheet = svod[str(month)]

        month_path = f"{static_path}\{MONTH_LIST[month - 1]}"

        for departement in os.listdir(month_path):
            try:
                file_name = list(filter(lambda x: "БИКУ" in x, os.listdir(f"{month_path}\{departement}")))[0]
                workbook = load_workbook(f'{month_path}\{departement}\{file_name}', read_only=True).worksheets[0]
                for row in workbook.iter_rows(min_row=8, values_only=True):
                    svod_sheet.append(row)
            except Exception as e:
                print(f"[INFO] Сформировать документ невозможно, Файл отсутсвует или не соответсвует синтаксису."
                      f"[INFO] Папка: {month_path}\{departement}")
                print(e)
        
        letter_list = ["R", "S", "T", "U", "V", "W"]
            
        for letter in letter_list:
            svod_sheet[f"{letter}4"] = f'=SUMIFS({letter}8:{letter}{svod_sheet.max_row + 1},$E8:$E{svod_sheet.max_row + 1},"Прием электроэнергии")'
            svod_sheet[f"{letter}5"] = f'=SUMIFS({letter}8:{letter}{svod_sheet.max_row + 1},$E8:$E{svod_sheet.max_row + 1},"Прием электроэнергии")'
            svod_sheet[f"{letter}6"] = f'={letter}4-{letter}5'
    
        svod.save(f"{static_path}\Сводная ведомость БИКУ.xlsx")


def calculate(month: str) -> None:
        static_file = load_workbook(f'{MAIN_DIR}\Бику\Реестр БИКУ.xlsx').worksheets[0]
        svod = load_workbook(f'{MAIN_DIR}\Сводный баланс\{datetime.now().year}\Бику\Сводная ведомость БИКУ.xlsx')
        svod_sheet  = svod.worksheets[month - 1]
        result = load_workbook('template\\bicu_v.xlsx')
        result_sheet = result.worksheets[0]
 
        print("[INFO] Собираем данные")

        for row_svod in range(8, svod_sheet.max_row + 1):
            for row_static in range(9, static_file.max_row + 1):
                if svod_sheet.cell(row=row_svod, column=3).value == static_file.cell(row=row_static, column=3).value and static_file.cell(row=row_static, column=48).value == "Активен":
                    result_sheet.append(
                        [
                            1, 
                            static_file.cell(row=row_svod, column=2).value, # B
                            static_file.cell(row=row_svod, column=3).value, # C
                            static_file.cell(row=row_svod, column=4).value, # D
                            static_file.cell(row=row_svod, column=5).value, # E
                            static_file.cell(row=row_svod, column=8).value, # F
                            static_file.cell(row=row_svod, column=9).value, # G
                            static_file.cell(row=row_svod, column=10).value, # H
                            static_file.cell(row=row_svod, column=13).value, # I
                            static_file.cell(row=row_svod, column=14).value, # J
                            static_file.cell(row=row_svod, column=15).value, # K
                            static_file.cell(row=row_svod, column=20).value, # L
                            static_file.cell(row=row_svod, column=21).value, # M
                            static_file.cell(row=row_svod, column=22).value, # N
                            static_file.cell(row=row_svod, column=23).value, # O
                            svod_sheet.cell(row=row_svod, column=17).value, # P
                            svod_sheet.cell(row=row_svod, column=17).value, # Q
                            "=Q{0}-P{0}".format(row_svod), # R
                            "=R{0}*O{0}".format(row_svod), # S
                            "", # T
                            static_file.cell(row=row_svod, column=46).value, # U
                            "", # V
                            "=S{0}+T{0}+U{0}+V{0}".format(row_svod), # W
                            static_file.cell(row=row_svod, column=27).value, # X
                            static_file.cell(row=row_svod, column=28).value, # Y
                            static_file.cell(row=row_svod, column=29).value, # Z
                            static_file.cell(row=row_svod, column=32).value, # AA
                            static_file.cell(row=row_svod, column=45).value, # AB
                            static_file.cell(row=row_svod, column=47).value # AC
                        ]
                    )
                    break
        print("[INFO] Данные загружены")
        result.save(MAIN_DIR + "\\result.xlsx")

        for departament in DEPARTAMENT_NAMES_WITH_KEYS.keys():
            departament_data = []
            for row in range(6, result_sheet.max_row +1):
                try: 
                    if result_sheet.cell(row=row, column=3).value[2:4] == departament:
                        departament_data.append([result_sheet.cell(row=row, column=i).value for i in range(1, result_sheet.max_column + 1)])
                except Exception as e:
                    pass
            workbook = load_workbook('./template/bicu_v.xlsx')
            worksheet = workbook.worksheets[0]
            for departament_row in departament_data:
                worksheet.append(departament_row)
            for report_row in range(6, worksheet.max_row+1):
                worksheet["R{0}".format(report_row)] = "=Q{0}-P{0}".format(report_row)
                worksheet["S{0}".format(report_row)] = "=R{0}*O{0}".format(report_row)
                worksheet["W{0}".format(report_row)] = "=S{0}+T{0}+U{0}+V{0}".format(report_row)
            
            
            letter_list = ["R", "S", "T", "U", "V", "W"]
            
            for letter in letter_list:
                worksheet[f"{letter}4"] = f'=SUMIFS({letter}8:{letter}{worksheet.max_row + 1},$E8:$E{worksheet.max_row + 1},"Прием электроэнергии")'
                worksheet[f"{letter}5"] = f'=SUMIFS({letter}8:{letter}{worksheet.max_row + 1},$E8:$E{worksheet.max_row + 1},"Прием электроэнергии")'
                worksheet[f"{letter}6"] = f'={letter}4-{letter}5'
            
        
            stuf_name = DEPARTAMENT_NAMES_WITH_KEYS[departament]
            mon = MONTH_LIST[month]

            workbook.save(f"{MAIN_DIR}\Сводный баланс\{datetime.now().year}\Бику\{MONTH_LIST[month]}\{stuf_name}\Ведомоть БИКУ {stuf_name} {mon} {datetime.now().year}.xlsx")
            print(f"[INFO] Файл " + f"Ведомоть БИКУ {stuf_name} {mon} {datetime.now().year}.xlsx сформирован")
