import os

from openpyxl import load_workbook


class Avarage:

    def get_data_from_files(self):
        for folder in os.listdir('../group'):
            hashtable = dict()
            for file in os.listdir(f'../group/{folder}'):
                workbook = load_workbook(f"../group/{folder}/{file}", data_only=True)
                worksheet = workbook.worksheets[0]
                for row in range(9, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=7).value
                    if cell is None:
                        continue
                    if cell not in hashtable.keys():
                        hashtable[cell] = []
                    if worksheet.cell(row=row, column=9).value != 0 or worksheet.cell(row=row, column=9).value != '0':
                        hashtable[cell].append(worksheet.cell(row=row, column=9).value)
            print(hashtable)

    def process_data(self):
        pass

    def add_add_in_file(self, data):
        workbook = load_workbook("./template/average.xlsx")
        print(workbook)
