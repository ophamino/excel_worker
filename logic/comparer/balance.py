from typing import Any
import os
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from logic.const import MAIN_DIR, MONTH_LIST


class Balance:
    """
    Класс для формирования сводного баланса
    """    
    
    
    def open_excel(self, path: str, data_only: bool = False, read_only: bool = False) -> Workbook:
        """
        Открывает файл, если его не существует - создаёт и открывает\n
        Args: path (str): путь к файлу, где должен находиться файл
        """
        if not os.path.exists(path):
            workbook = Workbook()
            workbook.save(path)
        return load_workbook(path, data_only=data_only, read_only=read_only)
    
    def create_months_sheet_if_not_exists(self, workbook: Workbook, month: str | int | None = None) -> None:
        """
       Создает месячный лист, если его не существует

        Args:
            workbook (Workbook): Документ excel, который будет проверяться\n\n
            month (str | int | None): Месячный лист, который должен находится в файле:\n
            Если значение `str`, то лист создаётся по названию;\n
            Если знаение `int`, то лист создается по номеру согласно порядку;\n
            Если значение `None`, то создается лист за текущий месяц.\n
            По умолчанию `None`.
        """

        sheetnames = workbook.sheetnames

        def create(workbook: Workbook, month: str | int, sheetnames: list[str]):
            if month not in sheetnames:
                workbook.create_sheet(month)

        if isinstance(month, str):
            if month in MONTH_LIST:
                create(workbook, month, sheetnames)
            else:
                raise ValueError(f"Указано неправильное название месяца: {month}")

        if isinstance(month, int):
            if 1 <= month <= len(MONTH_LIST):
                month = MONTH_LIST[month - 1]
                create(workbook, month, sheetnames)
            else:
                raise ValueError(f"Указан неправильный номер месяца: {month}")

        if month is None:
            month = datetime.now().month
            month = MONTH_LIST[month - 1]
            create(workbook, month, sheetnames)
        
            
    def open_sheet(self, workbook: Workbook, month: str | int) -> Worksheet:
        """
        Открывает лист по номеру или названию месяца

        Args:
            workbook (Workbook): Excel файл, лист которого нужно открыть\n
            month (str | int): Номер или название месяца

        Returns:
            Worksheet: Лист Excel файла
        """        
        self.create_months_sheet_if_not_exists(workbook, month)
        if isinstance(month, int):
            month = MONTH_LIST[month - 1]
        return workbook[month]


    def serialize_network(self, month: str | int) -> dict[str, dict[str, str]]:
        """
        Функция для формирования хэш-таблицы сети из файлы "Структура сети.xlsx"
        Args:
            month (str | int): Номер месяца

        Returns:
            dict[str, dict[str, str]]: Данные структуры сети
        """
        data = {}
        
        network_structure_path = f"{MAIN_DIR}\Структура сети\Свод ОЭСХ.xlsx"
        network_structure = self.open_excel(network_structure_path, data_only=True)
        network = self.open_sheet(network_structure, month)

        for row in range(2, network.max_row + 1):
            data[network.cell(row=row, column=1).value] = {
                "name": network.cell(row=row, column=2).value,
                "consumption": 0,
                "reception": 0,
                "transmission": 0,
                "balance": 0,
                "waste": 0,
                "foreign_key": network.cell(row=row, column=3).value,
                "foreign_key_name": network.cell(row=row, column=4).value,
            }

        return data
        
    
    def serialize_bicu(self, month: str | int) -> dict[str, dict[str, str]]:
        """
        Функция для формирования хэш-таблицы потребителей из файлы "Сводная ведомость БИКУ.xlsx"
        Args:
            month (str | int): Номер или название месяца

        Returns:
            dict[str, dict[str, str]]: Данные Бику
        """
        data = {}
        year = datetime.now().year
        
        bicu_path = f"{MAIN_DIR}\Сводный баланс\{year}\Бику\Сводная ведомость БИКУ.xlsx"
        bicu_file = self.open_excel(bicu_path, data_only=True)
        bicu = self.open_sheet(bicu_file, month)
        
        for row in range(8, bicu.max_row + 1):
            data[bicu.cell(row=row, column=3).value] = {
                "ID": bicu.cell(row=row, column=3).value,
                "name": bicu.cell(row=row, column=6).value,
                "status": bicu.cell(row=row, column=5).value,
                "expenses": bicu.cell(row=row, column=23).value,
                "foreign_key": bicu.cell(row=row, column=29).value,
            }

        return data
    
    
    def serialize_consumers(self, month: str | int) -> dict[str, dict[str, str]]:
        """
        Функция для формирования хэш-таблицы потребителей из файлы "Сводная ведомость.xlsx"
        Args:
            month (str | int): Номер или название месяца

        Returns:
            dict[str, dict[str, str]]: Данные потребителей
        """        
        data = {}
        year = datetime.now().year
        
        consumers_path = f"{MAIN_DIR}\Сводный баланс\{year}\УПП\Сводная ведомость потребителей.xlsx"
        consumers_file = self.open_excel(consumers_path, data_only=True)
        consumers = self.open_sheet(consumers_file, month)
        
        for row in range(6, consumers.max_row + 1):
            data[consumers.cell(row=row, column=3).value] = {
                "ID": consumers.cell(row=row, column=3).value,
                "name": consumers.cell(row=row, column=8).value,
                "expenses": consumers.cell(row=row, column=29).value,
                "foreign_key": consumers.cell(row=row, column=39).value,
            }

        return data
        

    def serialize_balance(self, month: str | int) -> dict[str, dict[str, str]]:
        """
        Функция для формирования данных сводного баланса из файлов "Структура сети.xlsx" и "Сводная ведомость.xlsx"

        Args:
            month (str | int): Номер или название месяцау

        Returns:
            dict[str, dict[str, str]]: Хэш-таблица баланса
        """
        consumers_data = self.serialize_consumers(month)
        network_data = self.serialize_network(month)
        bicu_data = self.serialize_bicu(month)
        
        balance = network_data

        for value in consumers_data.values():
            foreign_key = value["foreign_key"]
            if foreign_key:
                balance[foreign_key]["consumption"] += value["expenses"]

        for value in bicu_data.values():
            foreign_key = value["foreign_key"]
            if foreign_key:
                if value["status"] == "Прием электроэнергии":
                    balance[foreign_key]["reception"] += value["expenses"]
                if value["status"] == "Передача электроэнергии":
                    balance[foreign_key]["transmission"] += value["expenses"]
                if value["status"] not in ("Прием электроэнергии", "Передача электроэнергии"):
                    raise ValueError(
                        f'Неверные данные в Сводной ведомости БИКУ, - ID {value["ID"]} Должно быть "Прием электроэнергии" или "Передача электроэнергии"'
                        )
                    
        for value in network_data.values():
            foreign_key = value["foreign_key"]
            if foreign_key:
                balance[foreign_key]["consumption"] += value["consumption"]
                balance[foreign_key]["reception"] += value["reception"]
                balance[foreign_key]["transmission"] += value["transmission"]
                
        for key in balance.keys():
            balance[key]["balance"] = balance[key]["reception"] - balance[key]["transmission"]
            balance[key]["waste"] = balance[key]["balance"] - balance[key]["consumption"]

        return balance
        
    def create_balance(self, month: str | int) -> None:
        """
        Функция для вставки итоговых данных в файл "Сводный баланс"
        """
        year = datetime.now().year
        balance_path = f"{MAIN_DIR}\Сводный баланс\{year}\Сводный баланс.xlsx"
        balance_file = self.open_excel(balance_path, data_only=True)
        balance = self.open_sheet(balance_file, month)
        
        data = self.serialize_balance(month)
        
        balance.append(
            (
                "№", "Идентификатор", "Наименование", "Вход", "Выход", "Сальдо переток", "Полезный отпуск", "Потери"
            )
        )
        number = 1
        for key, value in data.items():
            balance.append(
                (
                    number, key, value["name"], value["reception"],
                    value["transmission"], value["balance"], value["consumption"], value["waste"]
                )
            )
            number = number + 1
        
        balance_file.save(balance_path)


    def __call__(self, *args: Any, **kwds: Any) -> Any: 
        pass
