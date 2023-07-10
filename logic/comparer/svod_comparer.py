from openpyxl import load_workbook

from datetime import datetime
import os

from logic.const import MAIN_DIR, MONTH_LIST

class Comparer:

    def __init__(self) -> None:
        self.main_dir = MAIN_DIR
        self.year = datetime.now().year
        self.month = datetime.now().month

    def comparer(self, file_status: str, month: str) -> None:
        static_path = f'{self.main_dir}\Сводный баланс\{self.year}\УПП'

        if not static_path in os.listdir(static_path):
            svod = load_workbook('template/svod.xlsx')
            svod.save(f"{static_path}\Сводная ведомость {file_status} потребления.xlsx")
        svod = load_workbook(f"{static_path}\Сводная ведомость {file_status} потребления.xlsx")
        svod_sheet = svod[str(month)]

        month_path = f"{static_path}\{MONTH_LIST[month - 1]}"
        print(month_path)

        for departement in os.listdir(month_path):
            try:
                file_name = list(filter(lambda x: file_status in x, os.listdir(f"{month_path}\{departement}")))[0]
                workbook = load_workbook(f'{month_path}\{departement}\{file_name}', read_only=True).worksheets[0]
                for row in workbook.iter_rows(min_row=11, values_only=True):
                    svod_sheet.append(row)
            except Exception as e:
                print(f"[INFO] Сформировать документ невозможно, Файл отсутсвует или не соответсвует синтаксису."
                      f"[INFO] Папка: {month_path}\{departement}")
                print(e)
        
        for report_row in range(6, svod_sheet.max_row+1):
            svod_sheet.cell(row=report_row, column=23).value = "=V{0}-U{0}".format(report_row)
            svod_sheet.cell(row=report_row, column=24).value = "=W{0}*T{0}".format(report_row)
            svod_sheet.cell(row=report_row, column=29).value = "=X{0}+Y{0}+Z{0}+AA{0}+AB{0}".format(report_row)
                
        svod.save(f"{static_path}\Сводная ведомость {file_status} потребления.xlsx")
    
    
    def total_comparer(self, month):
        static_path = f'{self.main_dir}\Сводный баланс\{self.year}\УПП'
        
        if not static_path in os.listdir(static_path):
            svod = load_workbook('template/svod.xlsx')
            svod.save(f"{static_path}\Сводная ведомость потребителей.xlsx")
        svod = load_workbook(f"{static_path}\Сводная ведомость потребителей.xlsx")
        svod_sheet = svod[str(month)]
        try:
            comerce = load_workbook(f"{static_path}\Сводная ведомость Коммерческого потребления.xlsx")[str(month)]
            individual = load_workbook(f"{static_path}\Сводная ведомость Бытового потребления.xlsx")[str(month)]
        except Exception as e:
            print(e)
        
        for row in individual.iter_rows(min_row=5, values_only=True):
            svod_sheet.append(row)
        for row in comerce.iter_rows(min_row=5, values_only=True):
            svod_sheet.append(row)
        svod.save(f"{static_path}\Сводная ведомость потребителей.xlsx")
