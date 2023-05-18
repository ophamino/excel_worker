from openpyxl import load_workbook


class Formater:
    """Класс для форматирования файлов"""

    def __init__(self) -> None:
        pass

    def add_numeric(self, file_name: str, skip_rows: int) -> None:
        """Функция для нумерации строк в файле"""
        workbook = load_workbook(file_name)
        sheets = workbook.sheetnames

        for sheet in sheets:
            work_sheet = workbook[sheet]
            for row in range(skip_rows, work_sheet.max_row + 1):
                work_sheet.cell(row=row, column=1).value = row - skip_rows
