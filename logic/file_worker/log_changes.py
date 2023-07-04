import os
from  datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from logic.const import MAIN_DIR


static_file_name = 'Реестр потребителей — копия.xlsx'
changes_file_name = 'Реестр потребителей для сравнения.xlsx'
log_file_name = 'Журнал изменений.xlsx'


time = datetime.now().time()


def get_status_data(identifiers: list, status: str) -> None:
    status_data = []
    if identifiers:
        for item in identifiers:
            status_data.append([1, item, '', '', '', '', status,   datetime.now().date(),  time])
    return status_data


def get_changes_data(static_file_sheet: Worksheet, changes_file_sheet: Worksheet, static_row: int, changes_row: int):
    changes = []
    for column in range(2, static_file_sheet.max_column + 1):
        if static_file_sheet.cell(row=static_row, column=column).value != changes_file_sheet.cell(row=changes_row, column=column).value:
            changes.append(
                [
                    1,
                    static_file_sheet.cell(row=changes_row, column=4).value,
                    static_file_sheet.cell(row=4, column=column).value,
                    changes_file_sheet.cell(row=changes_row, column=column).coordinate,
                    changes_file_sheet.cell(row=changes_row, column=column).value,
                    static_file_sheet.cell(row=static_row, column=column).value,
                    'Изменено',
                     datetime.now().date(),
                    time
                ]
            )

    return changes


def search_changes(static_file_sheet: Worksheet, changes_file_sheet: Worksheet) -> list:
    changes_data = []
    for row_static in range(9, static_file_sheet.max_row + 1):
        for row_changes in range(9, changes_file_sheet.max_row + 1):
            if static_file_sheet.cell(row=row_static, column=4).value == changes_file_sheet.cell(row=row_changes, column=4).value:
                change =  get_changes_data(static_file_sheet, changes_file_sheet, row_static, row_changes)
                changes_data = changes_data + change
                break
    return changes_data


def append_data_in_log(data: list[str]) -> None:
    if log_file_name not in os.listdir(f'{MAIN_DIR}\Потребители'):
        log_file = load_workbook("./template/log.xlsx").save(f'{MAIN_DIR}\Потребители\{log_file_name}')
    log_file = load_workbook(f'{MAIN_DIR}\Потребители\{log_file_name}')
    try:
        log_sheet = log_file[str( datetime.now().date().year)]

        for log_data in data:
            log_sheet.append(log_data)
        log_file.save(f'{MAIN_DIR}\Потребители\{log_file_name}')
    except Exception:
        print("[ERROR] Листа с текущим годом не существует, создайте его перед тем как начать работу")


def change_log():
    static_file = load_workbook(f'{MAIN_DIR}\Потребители\{ static_file_name}', data_only=True)
    changes_file = load_workbook(f"./template/{ changes_file_name}", data_only=True)

    static_file_sheet =  static_file.worksheets[0]
    changes_file_sheet =  changes_file.worksheets[0]


    changes_ID = set([changes_file_sheet.cell(row=i, column=4).value for i in range(9, changes_file_sheet.max_row + 1)])
    static_ID = set([static_file_sheet.cell(row=i, column=4).value for i in range(9, static_file_sheet.max_row + 1)])

    added_ID = list(static_ID.difference(changes_ID))
    deleted_ID = list(changes_ID.difference(static_ID))

    print("[INFO] Начинаем поиск изменений")
    changes_data =  search_changes(static_file_sheet, changes_file_sheet)
    added_data =  get_status_data(added_ID, status="Добавлено")
    deleted_data =  get_status_data(deleted_ID, status="Удалено")

    all_data = changes_data + added_data + deleted_data
    static_file.save(f"./template/Реестр потребителей для сравнения.xlsx")
    print("[INFO] записываем изменения в Журнал")
    append_data_in_log(all_data)
    print(
        f"\n[INFO] Измененно строк: {len(changes_data)}\n"
        f"[INFO] Добавленно строк: {len(added_data)}\n"
        f"[INFO] Удаленно строк: {len(deleted_data)}\n"
        f"[INFO] Всего строк: {len(all_data)}\n"
    )
